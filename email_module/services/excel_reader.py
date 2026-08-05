"""
excel_reader.py

Handles loading, calculating, and validating the uploaded Excel workbook.
"""

import os
import subprocess
import tempfile
import io
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:
    """
    Reads the generated audit workbook.
    """

    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook: Workbook | None = None

    def load(self) -> Workbook:
            """
            Load the workbook, forcing formula calculation via LibreOffice if available.
            """
            try:
                # Handle both physical file paths and io.BytesIO (from our Streamlit integration)
                if isinstance(self.file_path, io.BytesIO):
                    input_bytes = self.file_path.getvalue()
                else:
                    with open(self.file_path, "rb") as f:
                        input_bytes = f.read()

                with tempfile.TemporaryDirectory() as temp_dir:
                    input_path = os.path.join(temp_dir, "raw_workbook.xlsx")
                    
                    with open(input_path, "wb") as f:
                        f.write(input_bytes)

                    # 1. CREATE A DEDICATED OUTPUT DIRECTORY
                    out_dir = os.path.join(temp_dir, "calculated")
                    os.makedirs(out_dir)

                    # Force formula calculation using LibreOffice headless mode
                    try:
                        subprocess.run(
                            [
                                "libreoffice",
                                "--headless",
                                "--nologo",
                                "--nofirststartwizard",
                                "--convert-to", "xlsx",
                                "--outdir", out_dir,  # 2. POINT TO THE NEW DIRECTORY
                                input_path
                            ],
                            check=True,
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE
                        )
                    except FileNotFoundError:
                        pass
                    except subprocess.CalledProcessError as e:
                        print(f"LibreOffice calculation failed: {e}")

                    # 3. LOAD THE CALCULATED FILE FROM THE NEW DIRECTORY
                    calculated_path = os.path.join(out_dir, "raw_workbook.xlsx")
                    
                    # Now load it with data_only=True to grab the freshly cached values
                    self.workbook = load_workbook(
                        filename=calculated_path,
                        data_only=True
                    )

                return self.workbook

            except Exception as e:
                raise Exception(f"Unable to open workbook.\n{e}")

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """
        Return worksheet by name.
        """

        if self.workbook is None:
            raise Exception("Workbook not loaded.")

        if sheet_name not in self.workbook.sheetnames:
            raise Exception(
                f"Worksheet '{sheet_name}' not found."
            )

        return self.workbook[sheet_name]

    def get_checklist_sheet(self) -> Worksheet:
        """
        Returns Checklist worksheet.
        """

        return self.get_sheet("Checklist")

    def get_score_sheet(self) -> Worksheet:
        """
        Returns Score Parameters worksheet.
        """

        return self.get_sheet("Score Parameters")

    def get_sheet_names(self) -> list[str]:
        """
        Return all sheet names.
        """

        if self.workbook is None:
            raise Exception("Workbook not loaded.")

        return self.workbook.sheetnames

    def validate(self) -> bool:
        """
        Validate required worksheets exist.
        """

        required = [
            "Checklist",
            "Score Parameters"
        ]

        names = self.get_sheet_names()

        missing = []

        for sheet in required:

            if sheet not in names:
                missing.append(sheet)

        if missing:
            raise Exception(
                f"Workbook missing sheet(s): {', '.join(missing)}"
            )

        return True