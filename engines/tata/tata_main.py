import os
import json
import pandas as pd
import formulas 

from openpyxl import load_workbook

from engines.tata.report_utils import create_output_paths
from engines.tata.pdf_utils import extract_pdf_header, extract_evidence_pages, merge_pdfs, excel_to_pdf
from engines.tata.excel_utils import populate_headers, populate_checklist

def generate_report(
    audit_id,
    master_file,
    client,
    template_type,
    pdf_file,
    annexure_pdf=None
):

    with open("templates.json", "r", encoding="utf-8") as f:
        template_repository = json.load(f)

    template_file = template_repository[client][template_type]

    output_folder = "output"
    os.makedirs(output_folder, exist_ok=True)

    df = pd.read_excel(master_file, dtype=str, keep_default_na=False)
    df.columns = df.columns.str.strip()

    audit_df = df[df["Audit ID"].astype(str).str.strip() == audit_id]

    if audit_df.empty:
        raise Exception(f"Audit ID '{audit_id}' not found")

    print(f"Found {len(audit_df)} records")
    first_row = audit_df.iloc[0]

    agency_code = str(first_row["Agency Code"]).strip()
    agency_name = str(first_row["Agency Name"]).strip()

    paths = create_output_paths(agency_code, agency_name)

    generated_excel = paths["excel"]
    generated_pdf = paths["pdf"]
    evidence_pdf = paths["evidence"]
    final_report_pdf = paths["final"]

    pdf_data = extract_pdf_header(pdf_file)

    wb = load_workbook(template_file)
    checklist_sheet = wb.sheetnames[0]
    ws = wb[checklist_sheet]

    populate_headers(ws, first_row, pdf_data)
    populate_checklist(ws, audit_df)
    
    # --- PAGE SCALING FIX ---
    # Lock the width so the scoretable never spills onto a second page
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    wb.save(generated_excel)
    wb.close()

    print(generated_excel)
    print("Excel workbook created")

    # ---------------------------------------------------------
    # NEW LOGIC: Dynamic Formula Injection (Preserves Formatting)
    # ---------------------------------------------------------
    try:
        print("Dynamically evaluating formulas...")
        
        # 1. Let formulas engine calculate and write the raw values to a temp folder
        temp_dir = os.path.join(output_folder, "temp_calc")
        os.makedirs(temp_dir, exist_ok=True)
        
        xl_model = formulas.ExcelModel().loads(generated_excel).finish()
        xl_model.calculate()
        xl_model.write(dirpath=temp_dir)
        
        temp_calc_path = os.path.join(temp_dir, os.path.basename(generated_excel))
        
        # 2. Open the formatted workbook and the temporary calculated workbook
        formatted_wb = load_workbook(generated_excel)
        calculated_wb = load_workbook(temp_calc_path, data_only=True)
        
        # 3. Inject calculated values back into the beautiful formatted workbook
        for sheet_name in formatted_wb.sheetnames:
            if sheet_name in calculated_wb.sheetnames:
                formatted_ws = formatted_wb[sheet_name]
                calculated_ws = calculated_wb[sheet_name]
                
                for row in formatted_ws.iter_rows():
                    for cell in row:
                        # If the cell is a formula, replace it with the computed hard number
                        if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                            calc_val = calculated_ws[cell.coordinate].value
                            cell.value = calc_val
        
        # 4. Save the formatted workbook and clean up the temporary files
        formatted_wb.save(generated_excel)
        formatted_wb.close()
        calculated_wb.close()
        
        os.remove(temp_calc_path)
        os.rmdir(temp_dir)
        
        print("Formulas successfully calculated and injected!")
    except Exception as e:
        print(f"Warning: Dynamic calculation failed: {e}")
    # ---------------------------------------------------------

    excel_to_pdf(generated_excel, generated_pdf)
    print("PDF created")

    extract_evidence_pages(pdf_file, evidence_pdf)
    print("Evidence PDF created")

    if annexure_pdf:
        merge_pdfs(generated_pdf, evidence_pdf, annexure_pdf, final_report_pdf)
    else:
        merge_pdfs(generated_pdf, evidence_pdf, None, final_report_pdf)

    print("Final report created")

    return {
        "excel": generated_excel,
        "pdf": generated_pdf,
        "evidence": evidence_pdf,
        "final": final_report_pdf
    }